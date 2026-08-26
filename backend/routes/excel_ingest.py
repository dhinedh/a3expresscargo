import io
import openpyxl
from decimal import Decimal
# pyrefly: ignore [missing-import]
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
# pyrefly: ignore [missing-import]
from sqlalchemy.orm import Session

from database import get_db
from models import Shipment, Customer, ShipmentProduct, ShipmentCustomer, ShipmentSequence
from calculation_engine import recalculate_shipment

router = APIRouter(prefix="/api/v1/excel", tags=["Excel Ingestion"])


def safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    try:
        if isinstance(val, (int, float)):
            return float(val)
        val_str = str(val).replace(",", "").strip()
        if not val_str or val_str.startswith("="):
            return default
        return float(val_str)
    except Exception:
        return default


@router.post("/ingest")
async def ingest_excel_workbook(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel workbook (.xlsx or .xls)")

    contents = await file.read()
    try:
        wb_data = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        wb_formulas = openpyxl.load_workbook(io.BytesIO(contents), data_only=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to load Excel workbook: {str(e)}")

    sheet_names = wb_data.sheetnames

    # 1. Parse Shipping Instructions (SI sheet)
    usd_rate = 83.5
    lkr_inr_rate = 3.75
    profit_margin_pct = 15.0
    common_expenses_inr = 0.0
    common_expenses_lkr = 0.0
    port_expenses_lkr = 0.0

    if "SI" in sheet_names:
        ws_si = wb_data["SI"]
        lkr_inr_rate = safe_float(ws_si.cell(9, 3).value, 3.75) or 3.75 # C9
        usd_rate = safe_float(ws_si.cell(10, 3).value, 83.5) or 83.5    # C10
        profit_margin_pct = safe_float(ws_si.cell(11, 3).value, 15.0) or 15.0 # C11
        common_expenses_inr = safe_float(ws_si.cell(23, 5).value, 0.0) # E23
        port_expenses_lkr = safe_float(ws_si.cell(28, 5).value, 0.0)   # E28

    # 2. Get or Create Customers
    # Look for P_1, P_2 sheets or default customer
    customers = []
    cust1 = db.query(Customer).filter(Customer.code == "CUST-001").first()
    if not cust1:
        cust1 = Customer(
            name="CUSTOMER 1 (A3 COLOMBO)",
            code="CUST-001",
            country="Sri Lanka",
            address="Colombo, Sri Lanka"
        )
        db.add(cust1)
        db.flush()
    customers.append(cust1)

    if "P_2" in sheet_names:
        cust2 = db.query(Customer).filter(Customer.code == "CUST-002").first()
        if not cust2:
            cust2 = Customer(
                name="CUSTOMER 2 (COLOMBO RETAIL)",
                code="CUST-002",
                country="Sri Lanka",
                address="Colombo, Sri Lanka"
            )
            db.add(cust2)
            db.flush()
        customers.append(cust2)

    # 3. Create Shipment Record
    # Extract sequence number or generate next
    seq = db.query(ShipmentSequence).filter(ShipmentSequence.financial_year == "2026-27").first()
    if not seq:
        seq = ShipmentSequence(financial_year="2026-27", last_sequence=10)
        db.add(seq)
        db.flush()
    else:
        seq.last_sequence += 1

    shipment_no = f"AEC/{seq.last_sequence:02d}/2026-27"
    
    shipment = Shipment(
        shipment_no=shipment_no,
        sequence_number=seq.last_sequence,
        financial_year="2026-27",
        shipment_date="2026-07-27",
        status="CONFIGURED",
        usd_rate=Decimal(str(usd_rate)),
        lkr_inr_rate=Decimal(str(lkr_inr_rate)),
        profit_margin_pct=Decimal(str(profit_margin_pct)),
        common_expenses_inr=Decimal(str(common_expenses_inr)),
        common_expenses_lkr=Decimal(str(common_expenses_lkr)),
        port_expenses_lkr=Decimal(str(port_expenses_lkr)),
        freight_allocation_mode="WEIGHT",
        notes=f"Auto-imported from Excel: {file.filename}"
    )
    db.add(shipment)
    db.flush()

    # Link Customers to Shipment
    for cust in customers:
        sc = ShipmentCustomer(
            shipment_id=shipment.id,
            customer_id=cust.id,
            allocation_pct=Decimal(str(100.0 / len(customers)))
        )
        db.add(sc)

    # 4. Extract Products from ProductList / P_1 / P_2 / Invoice_India
    products_created = 0

    # Build product master dictionary from ProductList
    product_catalog = {}
    if "ProductList" in sheet_names:
        ws_pl = wb_data["ProductList"]
        for r in range(5, ws_pl.max_row + 1):
            name_val = ws_pl.cell(r, 10).value # Col J: Product Name
            hsn_val = str(ws_pl.cell(r, 6).value or "").strip() # Col F: HSN Code
            if not name_val or str(name_val).strip() == "":
                continue

            name_clean = str(name_val).strip()
            cost_inr = safe_float(ws_pl.cell(r, 16).value, 0.0) # Col P: Purchase price INR
            case_qty = safe_float(ws_pl.cell(r, 13).value, 1.0)  # Col M: No. of cases
            pkt_g = safe_float(ws_pl.cell(r, 12).value, 0.0)    # Col L: Packet size (g)
            net_kg = safe_float(ws_pl.cell(r, 19).value, 0.0)   # Col S: Net KG

            product_catalog[name_clean] = {
                "hsn_code": hsn_val,
                "cost_inr": cost_inr,
                "case_qty": case_qty,
                "pkt_g": pkt_g,
                "net_kg": net_kg,
                "row_idx": r
            }

    # Extract items from customer sheet P_1 (Customer 1)
    target_sheets = [("P_1", customers[0].id)]
    if len(customers) > 1 and "P_2" in sheet_names:
        target_sheets.append(("P_2", customers[1].id))

    for sheet_title, customer_id in target_sheets:
        ws = wb_data[sheet_title]
        ws_form = wb_formulas[sheet_title]

        for r in range(5, ws.max_row + 1):
            p_name = ws.cell(r, 7).value # Col G: Product Name
            if not p_name or str(p_name).strip() == "":
                continue
            
            p_name_clean = str(p_name).strip()
            total_units = safe_float(ws.cell(r, 10).value, 0.0) # Col J: TOTAL UNIT
            price_pkt = safe_float(ws.cell(r, 11).value, 0.0)   # Col K: PRICE/PKT
            discount = safe_float(ws.cell(r, 13).value, 0.0)    # Col M: DISCOUNT LKR
            set_price = safe_float(ws.cell(r, 15).value, 0.0)   # Col O: PRICE/PKT (SET)
            short_q = safe_float(ws.cell(r, 17).value, 0.0)     # Col Q: SHORT QTY
            cost_inr = safe_float(ws.cell(r, 6).value, 0.0)     # Col F: NETBUY

            # Get master info
            master_info = product_catalog.get(p_name_clean, {})
            hsn_code = master_info.get("hsn_code", "")
            pkt_g = master_info.get("pkt_g", 0.0)
            net_kg = master_info.get("net_kg", 0.0)

            # If total_units is zero, default to 1
            qty = total_units if total_units > 0 else 1.0
            
            # Master cost in INR
            raw_cost_inr = master_info.get("cost_inr", 0.0) or cost_inr
            # If raw_cost_inr is line total, compute unit price in INR
            unit_cost_inr = raw_cost_inr / qty if (qty > 0 and raw_cost_inr > 500) else raw_cost_inr

            sp = ShipmentProduct(
                shipment_id=shipment.id,
                customer_id=customer_id,
                product_name=p_name_clean,
                product_category="General",
                hsn_code=hsn_code,
                quantity=Decimal(str(qty)),
                pkt_size_g=Decimal(str(pkt_g)),
                net_weight_kg=Decimal(str(net_kg)),
                weight_val=Decimal(str(net_kg)),
                weight_unit="KG",
                unit="PCS",
                purchase_price=Decimal(str(round(unit_cost_inr, 4))),
                currency="INR",
                discount_lkr=Decimal(str(discount)),
                final_quotation_price=Decimal(str(price_pkt)) if price_pkt > 0 else Decimal("0.0"),
                set_price_lkr=Decimal(str(set_price)),
                short_qty=Decimal(str(short_q))
            )
            db.add(sp)
            products_created += 1

    db.flush()

    # Recalculate financial engine
    recalculate_shipment(db, shipment)

    return {
        "status": "SUCCESS",
        "message": f"Successfully ingested Excel workbook '{file.filename}'",
        "shipment_id": shipment.id,
        "shipment_no": shipment.shipment_no,
        "financial_year": shipment.financial_year,
        "products_imported": products_created,
        "exchange_rates": {
            "usd_rate": usd_rate,
            "lkr_inr_rate": lkr_inr_rate,
            "profit_margin_pct": profit_margin_pct
        }
    }
