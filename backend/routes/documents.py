import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse, Response
# pyrefly: ignore [missing-import]
from sqlalchemy.orm import Session
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from database import get_db
from models import (
    Shipment, Customer, ShipmentProduct, Vendor,
    ShipmentVendorAllocation, ShipmentCustomerRequirement, ShipmentVendorProformaItem
)

router = APIRouter(prefix="/api/v1/shipments/{shipment_id}/documents", tags=["Documents"])

def build_pdf_header(story, title: str, subtitle: str, shipment_no: str, date_str: str):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor('#0F172A'),
        fontName='Helvetica-Bold'
    )
    sub_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#475569')
    )

    header_table_data = [
        [
            Paragraph(f"<b>A3 EXPRESS LOGISTICS</b><br/>{title}", title_style),
            Paragraph(f"<b>Invoice / Ref No:</b> {shipment_no}<br/><b>Date:</b> {date_str}<br/><b>Status:</b> CONFIRMED", sub_style)
        ]
    ]
    t = Table(header_table_data, colWidths=[320, 200])
    t.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('ALIGN', (1,0), (1,0), 'RIGHT')
    ]))
    story.append(t)
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#3B82F6'), spaceAfter=15))


@router.get("/quotation")
def generate_customer_quotation(shipment_id: int, customer_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")
    cust = db.query(Customer).filter(Customer.id == customer_id).first()
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")

    products = [p for p in s.products if p.customer_id == customer_id]
    if not products:
        raise HTTPException(status_code=400, detail="No products allocated for this customer in shipment")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    build_pdf_header(story, f"CUSTOMER QUOTATION", f"Customer: {cust.name}", s.shipment_no, s.shipment_date or "")

    styles = getSampleStyleSheet()
    normal_style = styles['Normal']

    # Customer info box
    cust_info = [
        [Paragraph(f"<b>QUOTATION TO:</b><br/><b>{cust.name}</b> ({cust.code})<br/>{cust.address or ''}<br/>Phone: {cust.phone or 'N/A'} | Tax ID: {cust.tax_id or 'N/A'}", normal_style)]
    ]
    t_cust = Table(cust_info, colWidths=[520])
    t_cust.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F8FAFC')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#E2E8F0')),
        ('PADDING', (0,0), (-1,-1), 8)
    ]))
    story.append(t_cust)
    story.append(Spacer(1, 15))

    # Product table headers
    table_data = [
        ["#", "Product Name", "HSN Code", "Qty", "Unit", "Quoted Price (LKR)", "Total Amount (LKR)"]
    ]

    grand_total = 0.0
    for idx, p in enumerate(products, 1):
        q = float(p.quantity or 1.0)
        unit_price = float(p.final_quotation_price or p.suggested_price or 0.0)
        total_price = q * unit_price
        grand_total += total_price

        table_data.append([
            str(idx),
            p.product_name,
            p.hsn_code or "-",
            f"{q:g}",
            p.unit or "PCS",
            f"{unit_price:,.2f}",
            f"{total_price:,.2f}"
        ])

    table_data.append(["", "", "", "", "", "GRAND TOTAL:", f"LKR {grand_total:,.2f}"])

    t_prod = Table(table_data, colWidths=[30, 180, 75, 45, 45, 95, 95])
    t_prod.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#CBD5E1')),
        ('ALIGN', (3,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F1F5F9')),
        ('LINEABOVE', (0,-1), (-1,-1), 1.5, colors.HexColor('#1E293B'))
    ]))
    story.append(t_prod)

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Quotation_{cust.code}_{s.shipment_no.replace('/', '_')}.pdf"}
    )


@router.get("/indian-invoice")
def generate_indian_invoice(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    build_pdf_header(story, "INDIAN COMMERCIAL INVOICE (INR)", "Export Manifest & Commercial Invoice", s.shipment_no, s.shipment_date or "")

    table_data = [
        ["#", "Customer", "Product Name", "HSN Code", "Qty", "Unit Price (INR)", "Total Value (INR)"]
    ]

    cust_map = {c.id: c.name for c in db.query(Customer).all()}
    grand_total_inr = 0.0

    for idx, p in enumerate(s.products, 1):
        c_name = cust_map.get(p.customer_id, "Customer")
        q = float(p.quantity or 1.0)
        inr_price = float(p.indian_price or p.purchase_price or 0.0)
        total_inr = q * inr_price
        grand_total_inr += total_inr

        table_data.append([
            str(idx),
            c_name[:15],
            p.product_name,
            p.hsn_code or "-",
            f"{q:g}",
            f"Rs. {inr_price:,.2f}",
            f"Rs. {total_inr:,.2f}"
        ])

    table_data.append(["", "", "", "", "", "TOTAL INVOICE (INR):", f"INR {grand_total_inr:,.2f}"])

    t_prod = Table(table_data, colWidths=[25, 95, 160, 65, 40, 80, 95])
    t_prod.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#065F46')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#CBD5E1')),
        ('ALIGN', (4,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#ECFDF5'))
    ]))
    story.append(t_prod)

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Indian_Invoice_{s.shipment_no.replace('/', '_')}.pdf"}
    )


@router.get("/colombo-invoice")
def generate_colombo_invoice(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    build_pdf_header(story, "COLOMBO IMPORT INVOICE (LKR)", "Sri Lanka Customs Import Entry Invoice", s.shipment_no, s.shipment_date or "")

    table_data = [
        ["#", "Product Name", "HSN Code", "Qty", "Base LKR", "Duty LKR", "Total Cost LKR"]
    ]

    grand_cost_lkr = 0.0
    grand_duty_lkr = 0.0

    for idx, p in enumerate(s.products, 1):
        q = float(p.quantity or 1.0)
        base = float(p.base_price_lkr or 0.0) * q
        duty = float(p.calculated_duty_lkr or 0.0) * q
        cost = float(p.total_cost_lkr or 0.0) * q
        grand_cost_lkr += cost
        grand_duty_lkr += duty

        table_data.append([
            str(idx),
            p.product_name,
            p.hsn_code or "-",
            f"{q:g}",
            f"{base:,.2f}",
            f"{duty:,.2f}",
            f"{cost:,.2f}"
        ])

    table_data.append(["", "", "TOTALS:", "", "", f"LKR {grand_duty_lkr:,.2f}", f"LKR {grand_cost_lkr:,.2f}"])

    t_prod = Table(table_data, colWidths=[25, 175, 75, 45, 80, 80, 80])
    t_prod.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1E3A8A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#CBD5E1')),
        ('ALIGN', (3,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#EFF6FF'))
    ]))
    story.append(t_prod)

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Colombo_Invoice_{s.shipment_no.replace('/', '_')}.pdf"}
    )


@router.get("/packing-list")
def generate_packing_list(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    build_pdf_header(story, "SHIPMENT PACKING LIST", "Package & Weight Verification Details", s.shipment_no, s.shipment_date or "")

    table_data = [
        ["#", "Customer", "Product Name", "Quantity", "Weight", "Unit"]
    ]

    cust_map = {c.id: c.name for c in db.query(Customer).all()}
    total_qty = 0.0
    total_weight = 0.0

    for idx, p in enumerate(s.products, 1):
        c_name = cust_map.get(p.customer_id, "Customer")
        q = float(p.quantity or 1.0)
        w = float(p.weight_val or 0.0)
        total_qty += q
        total_weight += w * q

        table_data.append([
            str(idx),
            c_name[:20],
            p.product_name,
            f"{q:g}",
            f"{w:g} {p.weight_unit or 'KG'}",
            p.unit or "PCS"
        ])

    table_data.append(["", "", "TOTAL SHIPMENT WEIGHT:", f"{total_qty:g} Units", f"{total_weight:g} KG", ""])

    t_prod = Table(table_data, colWidths=[30, 130, 190, 60, 70, 40])
    t_prod.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#475569')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#CBD5E1')),
        ('ALIGN', (3,0), (4,-1), 'RIGHT'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F8FAFC'))
    ]))
    story.append(t_prod)

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Packing_List_{s.shipment_no.replace('/', '_')}.pdf"}
    )


@router.get("/duty-report")
def generate_duty_report(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    build_pdf_header(story, "CUSTOMS DUTY BREAKDOWN REPORT", "Itemized Tariff Rates & Calculated Duty", s.shipment_no, s.shipment_date or "")

    table_data = [
        ["#", "Product", "HSN Code", "Gen Duty", "VAT", "PAL", "CESS", "SSCL", "Duty / Unit (LKR)"]
    ]

    total_duty = 0.0
    for idx, p in enumerate(s.products, 1):
        duty = float(p.calculated_duty_lkr or 0.0)
        total_duty += duty * float(p.quantity or 1.0)

        table_data.append([
            str(idx),
            p.product_name[:20],
            p.hsn_code or "-",
            p.general_duty_rate or "0%",
            p.vat_rate or "0%",
            p.pal_rate or "0%",
            p.cess_rate or "0%",
            p.sscl_rate or "0%",
            f"{duty:,.2f}"
        ])

    table_data.append(["", "", "", "", "", "", "", "TOTAL DUTY:", f"LKR {total_duty:,.2f}"])

    t_prod = Table(table_data, colWidths=[25, 120, 65, 50, 45, 45, 45, 45, 80])
    t_prod.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7C3AED')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#CBD5E1')),
        ('ALIGN', (3,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F5F3FF'))
    ]))
    story.append(t_prod)

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Duty_Report_{s.shipment_no.replace('/', '_')}.pdf"}
    )


def build_perfect_invoice_sheet(ws, shipment_no, date_str, consignee_name, consignee_address, products, usd_rate=83.5, currency="USD", title="COMMERCIAL INVOICE"):
    ws.views.sheetView[0].showGridLines = True
    
    font_title = Font(name="Arial", size=14, bold=True)
    font_header_box = Font(name="Arial", size=9, bold=True, color="1E293B")
    font_bold = Font(name="Arial", size=9, bold=True)
    font_regular = Font(name="Arial", size=9)
    
    thin_side = Side(border_style="thin", color="000000")
    border_box = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_header_tbl = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_double_bottom = Border(bottom=Side(border_style="double", color="000000"), top=thin_side)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    fill_header = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    # Title Row 1
    ws.merge_cells("A1:K1")
    ws["A1"] = title
    ws["A1"].font = font_title
    ws["A1"].alignment = align_center

    # Box 1: Exporter vs Invoice Details (Rows 2 to 6)
    ws.merge_cells("A2:E2")
    ws["A2"] = "EXPORTER"
    ws["A2"].font = font_header_box
    ws["A2"].fill = fill_header

    ws["A3"] = "A3 EXPRESS CARGO,"
    ws["A4"] = "NO.20, SUNDRAM 3RD STREET, FIRST LINE, VYASARPADI,"
    ws["A5"] = "CHENNAI-600 039, INDIA"
    ws["A6"] = "BRANCH OFFICE: OLD NO.127, NEW NO.20, NEW STREET, MANNADY, CHENNAI - 600001."
    for r in range(3, 7):
        ws[f"A{r}"].font = font_regular

    ws.merge_cells("F2:I2")
    ws["F2"] = "Invoice No. & Date"
    ws["F2"].font = font_header_box
    ws["F2"].fill = fill_header
    ws.merge_cells("F3:I3")
    ws["F3"] = f"{shipment_no}   DT. {date_str}"
    ws["F3"].font = font_bold

    ws.merge_cells("J2:K2")
    ws["J2"] = "Exporter's Ref"
    ws["J2"].font = font_header_box
    ws["J2"].fill = fill_header
    ws.merge_cells("J3:K3")
    ws["J3"] = "AEC/EXP/2026"
    ws["J3"].font = font_regular

    ws.merge_cells("F4:I4")
    ws["F4"] = "Other Reference(s)"
    ws["F4"].font = font_header_box
    ws["F4"].fill = fill_header
    ws.merge_cells("F5:I5")
    ws["F5"] = "IEC NO: 0411012345"
    ws["F5"].font = font_regular

    ws.merge_cells("J4:K4")
    ws["J4"] = "B/L NO."
    ws["J4"].font = font_header_box
    ws["J4"].fill = fill_header

    # Box 2: Consignee vs Country & Buyer (Rows 7 to 10)
    ws.merge_cells("A7:E7")
    ws["A7"] = "CONSIGNEE"
    ws["A7"].font = font_header_box
    ws["A7"].fill = fill_header

    ws["A8"] = consignee_name or "NN BROTHER & HOLDINGS PVT LTD"
    ws["A8"].font = font_bold
    ws["A9"] = consignee_address or "NO.77/1/A. DEWALA ROAD, MAKOLA NORTH, MAKOLA, SRI LANKA."
    ws["A9"].font = font_regular

    ws.merge_cells("F7:I7")
    ws["F7"] = "Country of Origin and Goods"
    ws["F7"].font = font_header_box
    ws["F7"].fill = fill_header
    ws.merge_cells("F8:I8")
    ws["F8"] = "INDIA"
    ws["F8"].font = font_bold

    ws.merge_cells("J7:K7")
    ws["J7"] = "Country of Destination"
    ws["J7"].font = font_header_box
    ws["J7"].fill = fill_header
    ws.merge_cells("J8:K8")
    ws["J8"] = "SRI LANKA"
    ws["J8"].font = font_bold

    ws.merge_cells("F9:K9")
    ws["F9"] = "Buyer (if other than consignee)"
    ws["F9"].font = font_header_box
    ws["F9"].fill = fill_header
    ws.merge_cells("F10:K10")
    ws["F10"] = "PAN ASIA BANK PLC, DAM STREET BRANCH, COLOMBO 12 SRI LANKA"
    ws["F10"].font = font_regular

    # Box 3: Vessel / Port / Delivery terms (Rows 11 to 14)
    ws.merge_cells("A11:C11")
    ws["A11"] = "Vessel / Flight No."
    ws["A11"].font = font_header_box
    ws["A11"].fill = fill_header
    ws.merge_cells("A12:C12")
    ws["A12"] = "BY SEA"
    ws["A12"].font = font_bold

    ws.merge_cells("D11:E11")
    ws["D11"] = "Port of Loading"
    ws["D11"].font = font_header_box
    ws["D11"].fill = fill_header
    ws.merge_cells("D12:E12")
    ws["D12"] = "KATTUPALLI, INDIA"
    ws["D12"].font = font_bold

    ws.merge_cells("F11:K11")
    ws["F11"] = "Terms of Delivery and Payment"
    ws["F11"].font = font_header_box
    ws["F11"].fill = fill_header
    ws.merge_cells("F12:K12")
    ws["F12"] = "C&F COLOMBO - DP AT SIGHT"
    ws["F12"].font = font_bold

    ws.merge_cells("A13:C13")
    ws["A13"] = "Port of Discharge"
    ws["A13"].font = font_header_box
    ws["A13"].fill = fill_header
    ws.merge_cells("A14:C14")
    ws["A14"] = "COLOMBO, SRI LANKA"
    ws["A14"].font = font_bold

    ws.merge_cells("D13:K13")
    ws["D13"] = "Final Place of Delivery"
    ws["D13"].font = font_header_box
    ws["D13"].fill = fill_header
    ws.merge_cells("D14:K14")
    ws["D14"] = "COLOMBO, SRI LANKA"
    ws["D14"].font = font_bold

    # Apply Borders on Header Box grid
    for r in range(2, 15):
        for c in range(1, 12):
            cell = ws.cell(r, c)
            cell.border = border_box

    # Table Header Row 15 & 16
    ws.merge_cells("A15:A16")
    ws["A15"] = "Marks & Nos."
    
    ws.merge_cells("B15:E16")
    ws["B15"] = "Description of Goods"
    
    ws.merge_cells("F15:F16")
    ws["F15"] = "HS CODE"
    
    ws.merge_cells("G15:G16")
    ws["G15"] = "Quantity"
    
    ws.merge_cells("H15:H16")
    ws["H15"] = "Unit"
    
    ws.merge_cells("I15:I16")
    ws["I15"] = "No. of Pkgs"
    
    ws.merge_cells("J15:J16")
    ws["J15"] = f"Rate ({currency})"
    
    ws.merge_cells("K15:K16")
    ws["K15"] = f"Amount ({currency})"

    for r in range(15, 17):
        for c in range(1, 12):
            cell = ws.cell(r, c)
            cell.font = font_bold
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_header_tbl

    # Sub-header Row 17: Summary line
    ws.merge_cells("A17:A17")
    ws["A17"] = "VN/CMB"
    ws["A17"].alignment = align_center
    ws["A17"].font = font_bold
    
    ws.merge_cells("B17:E17")
    ws["B17"] = "AS PER ITEM DETAILS BELOW"
    ws["B17"].alignment = align_left
    ws["B17"].font = font_bold

    total_pkgs = len(products)
    ws["I17"] = f"{total_pkgs} PKGS"
    ws["I17"].font = font_bold
    ws["I17"].alignment = align_center

    for c in range(1, 12):
        ws.cell(17, c).border = border_header_tbl

    # Data Rows (Row 18 onwards)
    row_idx = 18
    total_amt = 0.0
    total_qty = 0.0

    for idx, p in enumerate(products, 1):
        q = float(getattr(p, "quantity", 1.0) or 1.0)
        p_name = getattr(p, "product_name", f"Product #{idx}")
        hsn = getattr(p, "hsn_code", "") or ""
        unit = getattr(p, "unit", "PCS") or "PCS"
        
        if currency == "LKR":
            rate = float(getattr(p, "final_quotation_price", 0.0) or getattr(p, "total_cost_lkr", 0.0) or 0.0)
        else:
            price_inr = float(getattr(p, "purchase_price", 0.0) or 0.0)
            rate = round(price_inr / usd_rate, 2) if usd_rate > 0 else 1.0

        amt = round(q * rate, 2)
        total_amt += amt
        total_qty += q

        ws.cell(row_idx, 1, idx).alignment = align_center
        
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)
        ws.cell(row_idx, 2, p_name).alignment = align_left
        
        ws.cell(row_idx, 6, hsn).alignment = align_center
        ws.cell(row_idx, 7, q).alignment = align_right
        ws.cell(row_idx, 7).number_format = "#,##0"
        
        ws.cell(row_idx, 8, unit).alignment = align_center
        ws.cell(row_idx, 9, 1).alignment = align_center
        
        num_fmt = "$#,##0.00" if currency == "USD" else "LKR #,##0.00"
        ws.cell(row_idx, 10, rate).alignment = align_right
        ws.cell(row_idx, 10).number_format = num_fmt
        
        ws.cell(row_idx, 11, amt).alignment = align_right
        ws.cell(row_idx, 11).number_format = num_fmt

        for c in range(1, 12):
            cell = ws.cell(row_idx, c)
            cell.font = font_regular
            cell.border = border_box

        row_idx += 1

    # Totals Row
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    ws.cell(row_idx, 1, f"TOTAL C&F {currency} :").alignment = align_right
    ws.cell(row_idx, 1).font = font_bold
    
    ws.cell(row_idx, 7, total_qty).alignment = align_right
    ws.cell(row_idx, 7).font = font_bold
    ws.cell(row_idx, 7).number_format = "#,##0"
    
    num_fmt = "$#,##0.00" if currency == "USD" else "LKR #,##0.00"
    ws.cell(row_idx, 11, total_amt).alignment = align_right
    ws.cell(row_idx, 11).font = font_bold
    ws.cell(row_idx, 11).number_format = num_fmt

    for c in range(1, 12):
        cell = ws.cell(row_idx, c)
        cell.border = border_double_bottom

    row_idx += 2

    # Declaration & Signature Block
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx+2, end_column=7)
    ws.cell(row_idx, 1, "Declaration:\nWe declare that this Invoice shows the actual price of the goods described and that all particulars are true and correct.").alignment = align_left
    ws.cell(row_idx, 1).font = font_regular

    ws.merge_cells(start_row=row_idx, start_column=8, end_row=row_idx+2, end_column=11)
    ws.cell(row_idx, 8, "For A3 EXPRESS CARGO\n\n\nAuthorized Signature & Date").alignment = align_center
    ws.cell(row_idx, 8).font = font_bold

    # Column Dimensions
    col_widths = {
        'A': 10, 'B': 16, 'C': 16, 'D': 16, 'E': 16,
        'F': 14, 'G': 12, 'H': 10, 'I': 12, 'J': 14, 'K': 16
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width


@router.get("/cmb-bank-excel")
def generate_cmb_bank_excel(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INVOICE"

    c_name = s.customers[0].customer.name if (s.customers and s.customers[0].customer) else "NN BROTHER & HOLDINGS PVT LTD"
    c_addr = s.customers[0].customer.address if (s.customers and s.customers[0].customer and s.customers[0].customer.address) else "NO.77/1/A. DEWALA ROAD, MAKOLA NORTH, MAKOLA, SRI LANKA."

    build_perfect_invoice_sheet(
        ws,
        shipment_no=s.shipment_no,
        date_str=s.shipment_date or "23.07.2026",
        consignee_name=c_name,
        consignee_address=c_addr,
        products=s.products,
        usd_rate=float(s.usd_rate or 83.5),
        currency="USD",
        title="COMMERCIAL INVOICE - BANK COPY"
    )

    buffer = io.BytesIO()
    wb.save(buffer)

    filename = f"A3_INVOICE_{s.shipment_no.replace('/', '_')}_CMB_BANK_DOCUMENT.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/indian-excel")
def generate_indian_excel(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "INVOICE"

    c_name = s.customers[0].customer.name if (s.customers and s.customers[0].customer) else "NN BROTHER & HOLDINGS PVT LTD"
    c_addr = s.customers[0].customer.address if (s.customers and s.customers[0].customer and s.customers[0].customer.address) else "NO.77/1/A. DEWALA ROAD, MAKOLA NORTH, MAKOLA, SRI LANKA."

    build_perfect_invoice_sheet(
        ws1,
        shipment_no=s.shipment_no,
        date_str=s.shipment_date or "23.07.2026",
        consignee_name=c_name,
        consignee_address=c_addr,
        products=s.products,
        usd_rate=float(s.usd_rate or 83.5),
        currency="USD",
        title="COMMERCIAL INVOICE"
    )

    buffer = io.BytesIO()
    wb.save(buffer)

    filename = f"A3_INDIAN_INVOICE_{s.shipment_no.replace('/', '_')}_CMB.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/coo-pdf")
def generate_coo_pdf(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    build_pdf_header(story, "CERTIFICATE OF ORIGIN (COO)", "Trade Agreement Origin Declaration (ISFTA/SAFTA)", s.shipment_no, s.shipment_date or "")

    table_data = [
        ["HS Code (4-Digit)", "Description of Goods", "Net Qty", "UOM", "Gross Qty", "FOB Value (USD)"]
    ]

    total_fob = 0.0
    for p in s.products:
        hsn_4 = (p.hsn_code or "")[:4] or "0000"
        q = float(p.quantity or 1.0)
        net_w = float(p.net_weight_kg or p.weight_val or 0.0)
        rate_usd = round(float(p.purchase_price or 0.0) / float(s.usd_rate or 83.5), 2) if float(s.usd_rate or 83.5) > 0 else 1.0
        fob_usd = round(q * rate_usd, 2)
        total_fob += fob_usd

        table_data.append([
            hsn_4,
            p.product_name[:35],
            f"{q:g}",
            p.unit or "PCS",
            f"{net_w:g} KG" if net_w > 0 else f"{q:g} PCS",
            f"${fob_usd:,.2f}"
        ])

    table_data.append(["", "TOTAL COO DECLARATION:", "", "", "", f"${total_fob:,.2f}"])

    t_prod = Table(table_data, colWidths=[100, 180, 55, 45, 65, 75])
    t_prod.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F766E')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 9),
        ('GRID', (0,0), (-1,-2), 0.5, colors.HexColor('#CBD5E1')),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#F0FDFA'))
    ]))
    story.append(t_prod)

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=Certificate_of_Origin_{s.shipment_no.replace('/', '_')}.pdf"}
    )


@router.get("/coo-excel")
def generate_coo_excel(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COOTemplate"
    ws.views.sheetView[0].showGridLines = True

    font_bold = Font(name="Calibri", size=9, bold=True)
    font_regular = Font(name="Calibri", size=9, bold=False)

    headers = [
        "Itchs Code", "Hs Code", "Description of the Export Products/ Goods",
        "Net Quantity", "Net Quantity UOM", "Item Invoice Value (In Invoice Currency)",
        "Marks & Number", "Number and Kind of Packages", "Gross Qty", "Gross Qty UOM",
        "Quantity to be printed in Coo", "Item FOB Value (In Invoice Currency)",
        "Item FOB Value (In INR)", "Item FOB Value (In USD)"
    ]

    for col_idx, h in enumerate(headers, 1):
        ws.cell(1, col_idx, h).font = font_bold

    row_idx = 2
    for p in s.products:
        hsn = p.hsn_code or "00000000"
        hsn_4 = hsn[:4]
        q = float(p.quantity or 1.0)
        net_w = float(p.net_weight_kg or p.weight_val or 0.0)
        price_inr = float(p.purchase_price or 0.0)
        usd_rate = float(s.usd_rate or 83.5)
        rate_usd = round(price_inr / usd_rate, 2) if usd_rate > 0 else 1.0
        fob_usd = round(q * rate_usd, 2)
        fob_inr = round(q * price_inr, 2)

        ws.cell(row_idx, 1, hsn).font = font_regular
        ws.cell(row_idx, 2, f"=LEFT(A{row_idx},4)").font = font_regular
        ws.cell(row_idx, 3, p.product_name).font = font_regular
        ws.cell(row_idx, 4, q).font = font_regular
        ws.cell(row_idx, 5, p.unit or "PCS").font = font_regular
        ws.cell(row_idx, 6, fob_usd).font = font_regular
        ws.cell(row_idx, 7, "AN/CMB").font = font_regular
        ws.cell(row_idx, 8, f"{len(s.products)} PKGS ONLY").font = font_regular
        ws.cell(row_idx, 9, net_w if net_w > 0 else q).font = font_regular
        ws.cell(row_idx, 10, "KILOGRAMS (KGS)").font = font_regular
        ws.cell(row_idx, 11, "Gross Quantity").font = font_regular
        ws.cell(row_idx, 12, fob_usd).font = font_regular
        ws.cell(row_idx, 13, fob_inr).font = font_regular
        ws.cell(row_idx, 14, fob_usd).font = font_regular
        row_idx += 1

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"COO_Template_{s.shipment_no.replace('/', '_')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/full-workbook")
def generate_full_workbook_excel(shipment_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")

    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    font_bold = Font(name="Calibri", size=9, bold=True)
    font_title = Font(name="Calibri", size=11, bold=True)
    font_regular = Font(name="Calibri", size=9, bold=False)

    # 1. DUTY_2025
    ws_duty = wb.create_sheet(title="DUTY_2025")
    ws_duty.append(["HS Code", "Description", "General Duty", "VAT", "PAL", "CESS", "SSCL", "Excise", "SCL"])
    for cell in ws_duty[1]:
        cell.font = font_bold
    for p in s.products:
        if p.hsn_code:
            ws_duty.append([p.hsn_code, p.product_name, p.general_duty_rate or "Free", p.vat_rate or "18%", p.pal_rate or "10%", p.cess_rate or "0%", p.sscl_rate or "2.5%", "", ""])

    # 2. SI
    ws_si = wb.create_sheet(title="SI")
    ws_si.append(["Parameter", "Value", "Notes"])
    ws_si.append(["INR/LKR Exchange Rate", float(s.lkr_inr_rate or 3.75), "Cell C9"])
    ws_si.append(["USD Exchange Rate", float(s.usd_rate or 83.5), "Cell C10"])
    ws_si.append(["Target Profit Margin %", float(s.profit_margin_pct or 15.0), "Cell C11"])
    ws_si.append(["Common Freight (INR)", float(s.common_expenses_inr or 0.0), "Cell E23"])
    ws_si.append(["Port Expenses (LKR)", float(s.port_expenses_lkr or 0.0), "Cell E28"])
    for cell in ws_si[1]:
        cell.font = font_bold

    # 3. InvoiceGen
    ws_igen = wb.create_sheet(title="InvoiceGen")
    ws_igen.append(["S.No", "Product Name", "HSN Code", "Quantity", "Unit", "Packet Size (g)", "Case Qty", "Net Weight (KG)"])
    for cell in ws_igen[1]:
        cell.font = font_bold
    for idx, p in enumerate(s.products, 1):
        ws_igen.append([idx, p.product_name, p.hsn_code or "", float(p.quantity or 1.0), p.unit or "PCS", float(p.pkt_size_g or 0.0), float(p.no_bags_qty or 1.0), float(p.net_weight_kg or p.weight_val or 0.0)])

    # 4. ProductList
    ws_pl = wb.create_sheet(title="ProductList")
    ws_pl.append(["S.No", "Product Name", "HSN Code", "NetBuy (INR)", "Base LKR", "Duty LKR", "Landed Cost LKR", "Suggested Price LKR", "FOB USD"])
    for cell in ws_pl[1]:
        cell.font = font_bold
    for idx, p in enumerate(s.products, 1):
        ws_pl.append([
            idx, p.product_name, p.hsn_code or "", float(p.purchase_price or 0.0),
            float(p.base_price_lkr or 0.0), float(p.calculated_duty_lkr or 0.0),
            float(p.total_cost_lkr or 0.0), float(p.suggested_price or 0.0),
            round(float(p.purchase_price or 0.0) / float(s.usd_rate or 83.5), 2) if float(s.usd_rate or 83.5) > 0 else 1.0
        ])

    # 5. Customer Sheets P_1 & P_2
    shipment_cust_list = [sc.customer_id for sc in s.customers] if s.customers else [1, 2]
    if len(shipment_cust_list) < 2:
        shipment_cust_list.extend([c.id for c in db.query(Customer).all() if c.id not in shipment_cust_list])
    if len(shipment_cust_list) < 2:
        shipment_cust_list = [1, 2]

    for c_idx, c_id in enumerate(shipment_cust_list[:2], 1):
        sheet_name = f"P_{c_idx}"
        ws_cust = wb.create_sheet(title=sheet_name)
        ws_cust.append(["S.No", "Material Code", "NetBuy (INR)", "Product Name", "Total Unit", "Price/Pkt (LKR)", "Discount (LKR)", "Set Price (LKR)", "Short Qty", "Net Settlement (LKR)", "Profit %"])
        for cell in ws_cust[1]:
            cell.font = font_bold

        cust_products = [p for p in s.products if p.customer_id == c_id]
        if not cust_products and c_idx == 1:
            cust_products = s.products

        for idx, p in enumerate(cust_products, 1):
            q = float(p.quantity or 1.0)
            set_p = float(p.set_price_lkr or p.final_quotation_price or 0.0)
            net_buy = float(p.purchase_price or 0.0) * float(s.lkr_inr_rate or 3.75) * q
            profit_pct = round(((float(p.net_settlement_lkr or 0.0) - net_buy) / net_buy * 100.0), 2) if net_buy > 0 else 0.0

            ws_cust.append([
                idx, f"MAT-{idx:04d}", float(p.purchase_price or 0.0), p.product_name,
                q, float(p.final_quotation_price or 0.0), float(p.discount_lkr or 0.0),
                set_p, float(p.short_qty or 0.0), float(p.net_settlement_lkr or 0.0), profit_pct
            ])

    # 6. Invoice_India
    ws_ii = wb.create_sheet(title="Invoice_India")
    c_name = s.customers[0].customer.name if (s.customers and s.customers[0].customer) else "NN BROTHER & HOLDINGS PVT LTD"
    c_addr = s.customers[0].customer.address if (s.customers and s.customers[0].customer and s.customers[0].customer.address) else "NO.77/1/A. DEWALA ROAD, MAKOLA NORTH, MAKOLA, SRI LANKA."
    build_perfect_invoice_sheet(
        ws_ii,
        shipment_no=s.shipment_no,
        date_str=s.shipment_date or "23.07.2026",
        consignee_name=c_name,
        consignee_address=c_addr,
        products=s.products,
        usd_rate=float(s.usd_rate or 83.5),
        currency="USD",
        title="INDIAN EXPORTER COMMERCIAL INVOICE"
    )

    # 7. Sheet2 (Category Aggregation)
    ws_s2 = wb.create_sheet(title="Sheet2")
    ws_s2.append(["Category", "Total Quantity", "Total Amount (USD)", "Avg Unit Rate (USD)"])
    for cell in ws_s2[1]:
        cell.font = font_bold
    categories = list({p.product_category or "General" for p in s.products})
    for cat in categories:
        cat_prods = [p for p in s.products if (p.product_category or "General") == cat]
        tot_q = sum(float(p.quantity or 1.0) for p in cat_prods)
        tot_usd = sum(float(p.quantity or 1.0) * (float(p.purchase_price or 0.0) / float(s.usd_rate or 83.5)) for p in cat_prods)
        avg_rate = round(tot_usd / tot_q, 3) if tot_q > 0 else 0.0
        ws_s2.append([cat, tot_q, round(tot_usd, 2), avg_rate])

    # 8. Sheet1 (HSN Pivot Summary)
    ws_s1 = wb.create_sheet(title="Sheet1")
    ws_s1.append(["HSN Code", "Count of Products", "Sum of USD", "Sum of Quantity", "Sum of Weight (KG)"])
    for cell in ws_s1[1]:
        cell.font = font_bold
    hsn_groups = {}
    for p in s.products:
        h = p.hsn_code or "Uncategorized"
        if h not in hsn_groups:
            hsn_groups[h] = []
        hsn_groups[h].append(p)
    for hsn, p_list in hsn_groups.items():
        cnt = len(p_list)
        q_sum = sum(float(p.quantity or 1.0) for p in p_list)
        w_sum = sum(float(p.net_weight_kg or p.weight_val or 0.0) * float(p.quantity or 1.0) for p in p_list)
        usd_sum = sum(float(p.quantity or 1.0) * (float(p.purchase_price or 0.0) / float(s.usd_rate or 83.5)) for p in p_list)
        ws_s1.append([hsn, cnt, round(usd_sum, 2), q_sum, round(w_sum, 2)])

    # 9. Invoice_Colombo
    ws_ic = wb.create_sheet(title="Invoice_Colombo")
    build_perfect_invoice_sheet(
        ws_ic,
        shipment_no=s.shipment_no,
        date_str=s.shipment_date or "23.07.2026",
        consignee_name=c_name,
        consignee_address=c_addr,
        products=s.products,
        usd_rate=float(s.usd_rate or 83.5),
        currency="LKR",
        title="COLOMBO IMPORTER COMMERCIAL INVOICE"
    )

    # 10. COOTemplate
    ws_coo = wb.create_sheet(title="COOTemplate")
    ws_coo.append(["Itchs Code", "Hs Code", "Description of Export Goods", "Net Quantity", "Net UOM", "FOB USD", "Marks & Number", "Packages", "Gross Qty", "Gross UOM"])
    for cell in ws_coo[1]:
        cell.font = font_bold
    for p in s.products:
        hsn = p.hsn_code or "00000000"
        q = float(p.quantity or 1.0)
        rate_usd = round(float(p.purchase_price or 0.0) / float(s.usd_rate or 83.5), 2) if float(s.usd_rate or 83.5) > 0 else 1.0
        ws_coo.append([hsn, f"=LEFT(A{ws_coo.max_row+1},4)", p.product_name, q, p.unit or "PCS", round(q * rate_usd, 2), "AN/CMB", f"{len(s.products)} PKGS ONLY", float(p.net_weight_kg or p.weight_val or 0.0), "KILOGRAMS (KGS)"])

    # Auto-adjust column widths across all sheets
    for sheet in wb.worksheets:
        for col in sheet.columns:
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    filename = f"A3EXPRESS_FULL_WORKBOOK_{s.shipment_no.replace('/', '_')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ─── Per-Vendor RFQ Document Endpoints (Requirement 9 & 10) ──────────────────────────

@router.get("/vendor-rfq/{vendor_id}/pdf")
def generate_vendor_rfq_pdf(shipment_id: int, vendor_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")
    v = db.query(Vendor).filter(Vendor.id == vendor_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="Vendor not found")

    # Fetch allocated requirement products for vendor
    allocs = db.query(ShipmentVendorAllocation).filter(
        ShipmentVendorAllocation.shipment_id == shipment_id,
        ShipmentVendorAllocation.vendor_id == vendor_id
    ).all()

    req_ids = [a.requirement_id for a in allocs if a.requirement_id]
    reqs = db.query(ShipmentCustomerRequirement).filter(
        ShipmentCustomerRequirement.id.in_(req_ids)
    ).all() if req_ids else []

    # Also fetch proforma items if already recorded
    pis = db.query(ShipmentVendorProformaItem).filter(
        ShipmentVendorProformaItem.shipment_id == shipment_id,
        ShipmentVendorProformaItem.vendor_id == vendor_id
    ).all()

    # Fallback to all requirements in shipment if no specific vendor allocation row yet
    if not reqs and not pis:
        reqs = db.query(ShipmentCustomerRequirement).filter(
            ShipmentCustomerRequirement.shipment_id == shipment_id
        ).all()

    # Fallback to all customer requirements in database if shipment requirements empty
    if not reqs and not pis and not s.products:
        reqs = db.query(ShipmentCustomerRequirement).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    story = []

    build_pdf_header(story, "REQUEST FOR QUOTATION (RFQ)", f"Supplier: {v.name} ({v.code})", s.shipment_no, s.shipment_date or "")

    styles = getSampleStyleSheet()
    normal_style = styles['Normal']

    # Vendor Info Box
    v_info = [
        [Paragraph(f"<b>TO SUPPLIER / VENDOR:</b><br/><b>{v.name}</b> ({v.code})<br/>Contact: {v.contact_person or 'N/A'} | Phone: {v.phone or 'N/A'}<br/>GSTIN: {v.gstin or 'N/A'} | Email: {v.email or 'N/A'}<br/>Address: {v.address or 'N/A'}", normal_style)]
    ]
    t_v = Table(v_info, colWidths=[520])
    t_v.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F1F5F9')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#CBD5E1')),
        ('PADDING', (0,0), (-1,-1), 8)
    ]))
    story.append(t_v)
    story.append(Spacer(1, 15))

    # Items table
    table_data = [
        ["#", "Product Name", "HSN", "Qty", "Units/Ctn", "Target Price", "MRP", "Disc %", "GST %", "Total Payable"]
    ]

    items_count = 0
    if pis:
        for idx, p in enumerate(pis, 1):
            items_count += 1
            table_data.append([
                str(idx),
                p.product_name,
                p.hsn_code or "-",
                f"{p.proforma_qty:g}",
                str(p.units_per_carton or 12),
                f"Rs.{p.proforma_price:.2f}",
                f"Rs.{p.mrp:.2f}" if p.mrp else "-",
                f"{p.discount_pct}%" if p.discount_pct else "0%",
                f"{p.gst_pct}%" if p.gst_pct else "18%",
                f"Rs.{p.total_payable:.2f}" if p.total_payable else f"Rs.{(p.proforma_qty * p.proforma_price):.2f}"
            ])
    elif reqs:
        for idx, r in enumerate(reqs, 1):
            items_count += 1
            table_data.append([
                str(idx),
                r.product_name,
                r.hsn_code or "-",
                f"{r.required_quantity:g}",
                "12",
                f"Rs.{r.target_price:.2f}" if r.target_price else "Rs.45.00",
                "Rs.60.00",
                "5%",
                "18%",
                f"Rs.{(r.required_quantity * (r.target_price or 45.0)):.2f}"
            ])
    else:
        ship_prods = s.products
        for idx, sp in enumerate(ship_prods, 1):
            items_count += 1
            table_data.append([
                str(idx),
                sp.product_name,
                sp.hsn_code or "-",
                f"{sp.quantity:g}",
                "12",
                f"Rs.{sp.purchase_price:.2f}" if sp.purchase_price else "Rs.45.00",
                "-",
                "0%",
                "18%",
                f"Rs.{(sp.quantity * (sp.purchase_price or 45.0)):.2f}"
            ])

    if items_count == 0:
        table_data.append(["1", "Ragi (Finger Millet)", "1008.2910", "12", "12", "Rs.45.00", "Rs.60.00", "5%", "18%", "Rs.540.00"])

    t_prod = Table(table_data, colWidths=[25, 140, 55, 35, 45, 55, 45, 35, 35, 50])
    t_prod.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F172A')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 8),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN', (2,0), (-1,-1), 'CENTER'),
    ]))
    story.append(t_prod)
    story.append(Spacer(1, 20))

    # Notes section
    story.append(Paragraph("<b>IMPORTANT INSTRUCTIONS FOR VENDOR:</b>", styles['Heading3']))
    story.append(Paragraph("1. Reply with Proforma Invoice (PI) detailing exact carton packaging specs (Cartons Count, Units per Carton, Unit Weight KG, Net & Gross Weights).", normal_style))
    story.append(Paragraph("2. Specify payment terms, dispatch lead time, and base unit prices in INR / USD.", normal_style))

    doc.build(story)
    buffer.seek(0)
    filename = f"RFQ_{v.code}_{s.shipment_no.replace('/', '_')}.pdf"
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"inline; filename={filename}"}
    )


@router.get("/vendor-rfq/{vendor_id}/excel")
def generate_vendor_rfq_excel(shipment_id: int, vendor_id: int, db: Session = Depends(get_db)):
    s = db.query(Shipment).filter(Shipment.id == shipment_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Shipment not found")
    v = db.query(Vendor).filter(Vendor.id == vendor_id).first()
    if not v:
        raise HTTPException(status_code=404, detail="Vendor not found")

    allocs = db.query(ShipmentVendorAllocation).filter(
        ShipmentVendorAllocation.shipment_id == shipment_id,
        ShipmentVendorAllocation.vendor_id == vendor_id
    ).all()

    req_ids = [a.requirement_id for a in allocs if a.requirement_id]
    reqs = db.query(ShipmentCustomerRequirement).filter(
        ShipmentCustomerRequirement.id.in_(req_ids)
    ).all() if req_ids else []

    pis = db.query(ShipmentVendorProformaItem).filter(
        ShipmentVendorProformaItem.shipment_id == shipment_id,
        ShipmentVendorProformaItem.vendor_id == vendor_id
    ).all()

    if not reqs and not pis:
        reqs = db.query(ShipmentCustomerRequirement).filter(
            ShipmentCustomerRequirement.shipment_id == shipment_id
        ).all()

    if not reqs and not pis and not s.products:
        reqs = db.query(ShipmentCustomerRequirement).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"RFQ-{v.code}"

    # Title
    ws.append(["AEC / A3 EXPRESS SOFTWARE — VENDOR PRODUCT REQUIREMENT SHEET"])
    ws.append([f"Shipment ID: {s.shipment_no}", f"Supplier Name: {v.name} ({v.code})", f"Date: {s.shipment_date or ''}"])
    ws.append([])

    headers = [
        "S.No", "Product Name", "HSN Code", "Required Quantity", "Unit",
        "Units Per Carton", "Unit Price (INR)", "MRP (INR)", "Discount %", "GST %", "Total Payable (INR)", "Notes / Remarks"
    ]
    ws.append(headers)

    font_bold = Font(bold=True)
    for cell in ws[4]:
        cell.font = font_bold
        cell.fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    if pis:
        for idx, p in enumerate(pis, 1):
            ws.append([
                idx, p.product_name, p.hsn_code or "", float(p.proforma_qty), "PCS",
                p.units_per_carton or 12, float(p.proforma_price or 0.0), float(p.mrp or 0.0), float(p.discount_pct or 0.0), float(p.gst_pct or 18.0), float(p.total_payable or 0.0), p.notes or ""
            ])
    elif reqs:
        for idx, r in enumerate(reqs, 1):
            ws.append([
                idx, r.product_name, r.hsn_code or "", float(r.required_quantity), r.unit or "PCS",
                12, float(r.target_price or 45.0), 60.0, 5.0, 18.0, float((r.required_quantity * (r.target_price or 45.0))), r.notes or ""
            ])
    else:
        ship_prods = s.products
        for idx, sp in enumerate(ship_prods, 1):
            ws.append([
                idx, sp.product_name, sp.hsn_code or "", float(sp.quantity), sp.unit or "PCS",
                12, float(sp.purchase_price or 45.0), 0.0, 0.0, 18.0, float((sp.quantity * (sp.purchase_price or 45.0))), ""
            ])

    buffer = io.BytesIO()
    wb.save(buffer)
    filename = f"RFQ_{v.code}_{s.shipment_no.replace('/', '_')}.xlsx"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


